from Books import Book
import openpyxl

FILENAME = "Records.xlsx"
WB = openpyxl.load_workbook(FILENAME)
BOOKS_SHEET = WB.active


def create_book():
    max_row = BOOKS_SHEET.max_row
    title = input("Enter book title:")
    author = input("Enter book author:")
    year = input("Enter book year:")
    new_book = Book(max_row+1, title, author, year)
    new_book.record(BOOKS_SHEET)
    print("Book recorded successfully.")


def rundown():
    count = 1
    for row in BOOKS_SHEET.rows:
        print(f"{count}  Book: {row[0].value}, Author: {row[1].value}, Year: {row[2].value}")
        count += 1


def commands_help():
    print('\n')
    print("Commands:")
    print("\"new\": to record a new book.")
    print("\"rundown\": to displays a list of all books.")
    print("\"close\": to close application.")
    print("\"delete {number of book}\": to delete a book.")
    print("\"rewrite {number of book}\": to rewrite a book\'s details.")
    print("\"edit title {number of book}\": to edit a book\'s title.")
    print("\"edit author {number of book}\": to edit a book\'s author.")
    print("\"edit year {number of book}\": to edit a book\'s year.")
    print("======================================================\n")


def loop():
    while True:
        user_input = input()

        if user_input == "new":
            create_book()
            WB.save(FILENAME)

        elif user_input == "close":
            break

        elif user_input.startswith("delete"):
            subject = user_input.replace("delete ", "")
            if subject.isnumeric():
                if 0 < int(subject) <= BOOKS_SHEET.max_row:
                    book_num = int(subject)
                    BOOKS_SHEET.delete_rows(book_num)
                    WB.save(FILENAME)
                    print("Deleted successfully")
                else:
                    print(f"Enter a number between 0 and {BOOKS_SHEET.max_row}")
            else:
                print("Please enter a number.")

        elif user_input == "rundown":
            rundown()

        elif user_input.startswith("rewrite"):
            subject = user_input.replace("rewrite ", "")
            if subject.isnumeric():
                if 0 < int(subject) <= BOOKS_SHEET.max_row:
                    order_input = int(subject)
                    input_title = input("Enter new title: ")
                    input_author = input("Enter new author: ")
                    input_year = input("Enter new year: ")
                    Book.rewrite(Book, BOOKS_SHEET, order_input, input_title, input_author, input_year)
                    WB.save(FILENAME)
                    print(f"Book number {order_input} edited.")
                else:
                    print(f"Enter a number between 0 and {BOOKS_SHEET.max_row}")
            else:
                print("Please enter a number.")

        elif user_input.startswith("edit title"):
            subject = user_input.replace("edit title ", "")
            if subject.isnumeric():
                if 0 < int(subject) <= BOOKS_SHEET.max_row:
                    order_input = int(subject)
                    input_title = input("Enter new title: ")
                    Book.edit_title(Book, BOOKS_SHEET, order_input, input_title)
                    WB.save(FILENAME)
                    print(f"Changed book title number {order_input} to \"{input_title}\"")
                else:
                    print(f"Enter a number between 0 and {BOOKS_SHEET.max_row}")
            else:
                print("Please enter a number.")

        elif user_input.startswith("edit author"):
            subject = user_input.replace("edit author ", "")
            if subject.isnumeric():
                if 0 < int(subject) <= BOOKS_SHEET.max_row:
                    order_input = int(subject)
                    input_author = input("Enter new author: ")
                    Book.edit_author(Book, BOOKS_SHEET, order_input, input_author)
                    WB.save(FILENAME)
                    print(f"Changed book author number {order_input} to \"{input_author}\"")
                else:
                    print(f"Enter a number between 0 and {BOOKS_SHEET.max_row}")
            else:
                print("Please enter a number.")

        elif user_input.startswith("edit year"):
            subject = user_input.replace("edit year ", "")
            if subject.isnumeric():
                if 0 < int(subject) <= BOOKS_SHEET.max_row:
                    order_input = int(subject)
                    input_year = input("Enter new year: ")
                    Book.edit_year(Book, BOOKS_SHEET, order_input, input_year)
                    WB.save(FILENAME)
                    print(f"Changed book year number {order_input} to \"{input_year}\"")
                else:
                    print(f"Enter a number between 0 and {BOOKS_SHEET.max_row}")
            else:
                print("Please enter a number.")

        else:
            print("Unknown command.")
